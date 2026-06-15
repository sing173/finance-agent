"""Excel builder for bank statement transactions using openpyxl"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Optional

from ..models import Transaction


class ExcelBuilder:
    """交易流水 Excel 导出器"""

    # ------------------------------------------------------------------ #
    #  原有方法：导出原始流水明细                                             #
    # ------------------------------------------------------------------ #

    def build(self, transactions: List[Transaction], output_path: str) -> str:
        """导出交易列表到 Excel（原始流水明细格式）"""
        wb = openpyxl.Workbook()

        # 交易明细
        ws = wb.active
        ws.title = '交易明细'
        ws['A1'] = '银行流水明细'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:I1')

        # 统计
        income_count = sum(1 for t in transactions if t.direction == 'income')
        expense_count = sum(1 for t in transactions if t.direction == 'expense')
        total_income = sum(t.amount for t in transactions if t.direction == 'income')
        total_expense = sum(t.amount for t in transactions if t.direction == 'expense')

        stats = [
            ('总交易数', len(transactions)),
            ('收入笔数', income_count),
            ('支出笔数', expense_count),
            ('收入合计', float(total_income)),
            ('支出合计', float(total_expense)),
        ]
        ws['A3'] = '统计摘要'
        ws['A3'].font = Font(bold=True, size=12)
        for i, (label, value) in enumerate(stats, start=4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

        # 交易列表
        data_start = len(stats) + 5
        headers = ['日期', '描述', '金额', '币种', '方向', '对方户名', '流水号', '本方帐号', '本方户名']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=data_start, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for row, t in enumerate(transactions, start=data_start + 1):
            ws.cell(row=row, column=1, value=t.date.isoformat())
            ws.cell(row=row, column=2, value=t.description)
            ws.cell(row=row, column=3, value=float(t.amount))
            ws.cell(row=row, column=4, value=t.currency)
            ws.cell(row=row, column=5, value=t.direction)
            ws.cell(row=row, column=6, value=t.counterparty)
            ws.cell(row=row, column=7, value=t.reference_number)
            ws.cell(row=row, column=8, value=t.account_number)
            ws.cell(row=row, column=9, value=t.account_name)

        ws.freeze_panes = f'A{data_start + 1}'
        self._auto_width(ws)
        wb.save(output_path)
        return output_path

    def _auto_width(self, ws):
        """自动调整列宽"""
        for col in ws.columns:
            column_letter = get_column_letter(col[0].column)
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # ------------------------------------------------------------------ #
    #  凭证 Excel 导出（v0.3.0 凭证系统：直接从分录 dict 列表写入）       #
    # ------------------------------------------------------------------ #

    # 金蝶凭证导出：PipelineEntry 字段 → (Excel 列号, 转换函数)
    # 声明式映射：新增字段只改这里
    VOUCHER_COLUMN_MAP: dict[int, tuple[str, Optional[str]]] = {
        1:  ('date', '_as_str'),
        3:  ('voucher_no', None),
        5:  ('entry_seq', None),
        6:  ('summary', None),
        7:  ('subject_code', None),
        8:  ('subject_name', None),
        9:  ('debit_amount', '_fmt_money'),
        10: ('credit_amount', '_fmt_money'),
        15: ('aux_category', '_or_none'),
        23: ('original_amount', '_fmt_orig'),
    }

    @staticmethod
    def _as_str(value):
        """转为字符串（None → ''）。"""
        return str(value) if value is not None else ''

    @staticmethod
    def _fmt_money(value):
        """格式化金额（#，##0.00）。"""
        if value is None:
            return None
        return float(value)

    @staticmethod
    def _fmt_orig(value):
        """格式化原币金额（##.00#####）。"""
        if value is None:
            return None
        return float(value)

    @staticmethod
    def _or_none(value):
        """空字符串转 None，否则原样返回。"""
        return value or None

    def build_voucher_from_entries(
        self,
        entries: list[dict],
        output_path: str = 'voucher.xlsx',
        period: str = '',
    ) -> str:
        """直接从分录 dict 列表导出凭证 Excel —— 不做科目重新匹配。

        用于 export 保持与预览完全一致。
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        sheet_name = f'凭证列表#{period}' if period else '凭证列表'
        ws.title = sheet_name

        HEADERS = [
            '日期', '凭证字', '凭证号', '附件数', '分录序号',
            '摘要', '科目代码', '科目名称', '借方金额', '贷方金额',
            '客户', '供应商', '职员', '项目', '部门', '存货',
            '自定义辅助核算类别', '自定义辅助核算编码',
            '自定义辅助核算类别1', '自定义辅助核算编码1',
            '数量', '单价', '原币金额', '币别', '汇率',
        ]

        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for i in range(1, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15

        for row_idx, e in enumerate(entries, start=2):
            # 声明式映射驱动写入
            for col, (field, transform_name) in self.VOUCHER_COLUMN_MAP.items():
                value = e.get(field)
                if transform_name:
                    transform = getattr(self, transform_name)
                    value = transform(value)
                cell = ws.cell(row=row_idx, column=col, value=value)
                # 金额列加 number_format
                if transform_name == '_fmt_money':
                    cell.number_format = '#,##0.00'
                elif transform_name == '_fmt_orig':
                    cell.number_format = '##.00#####'

            # 常量列
            ws.cell(row=row_idx, column=2, value='记')
            ws.cell(row=row_idx, column=4, value=0)
            ws.cell(row=row_idx, column=24, value='RMB')
            cell_rate = ws.cell(row=row_idx, column=25, value=1.0)
            cell_rate.number_format = '##.00#####'

        ws.freeze_panes = 'A2'
        wb.save(output_path)
        return output_path
