"""招商银行交易流水 Excel 解析器。

处理招商银行网银导出的交易流水 Excel (.xlsx) 文件。

Excel 格式：
- 前几行为元数据（账号、期初/期末余额等）
- 标题行包含：账号、账号名称、币种、交易日期、起息日、交易类型、
  借方发生额、贷方发生额、余额、摘要、流水号、收(付)方单位名称等
- 标题行之后为数据行

列映射：
- 起息日 → date
- 借方发生额 → expense amount
- 贷方发生额 → income amount
- 摘要 → description
- 流水号 → reference_number
- 收(付)方单位名称 → counterparty
- 收(付)方账号 → counterparty account (in notes)
- 余额 → balance (in notes)
"""

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple

from ..models import ParseResult, Transaction


class CMBExcelParser:
    """招商银行交易流水 Excel 解析器"""

    BANK_NAME = "招商银行"

    # 标题行标识关键字（用于定位标题行）
    HEADER_KEYWORDS = ["起息日", "借方金额", "贷方金额"]

    # 列名 → 标准化字段名
    COLUMN_MAP = {
        "起息日": "value_date",
        "交易日": "trade_date",
        "交易日期": "trade_date",
        "借方金额": "debit_amount",
        "贷方金额": "credit_amount",
        "余额": "balance",
        "摘要": "summary",
        "流水号": "serial_no",
        "收(付)方单位名称": "counterparty",
        "收(付)方名称": "counterparty",
        "收(付)方账号": "counterparty_account",
        "账号": "account_no",
        "账号名称": "account_name",
        "交易类型": "transaction_type",
        "业务参考号": "business_ref",
        "扩展摘要": "extended_summary",
    }

    def __init__(self):
        self.confidence = 1.0

    def parse(self, file_path: str) -> ParseResult:
        import openpyxl

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return ParseResult(
                transactions=[], bank=self.BANK_NAME,
                confidence=0, errors=[f"Excel 打开失败: {e}"],
            )

        ws = wb.active
        if ws is None or ws.max_row < 2:
            wb.close()
            return ParseResult(
                transactions=[], bank=self.BANK_NAME,
                confidence=0, errors=["Excel 无数据"],
            )

        # 定位标题行
        header_row, col_map = self._find_header(ws)
        if header_row is None:
            wb.close()
            return ParseResult(
                transactions=[], bank=self.BANK_NAME,
                confidence=0, errors=["未找到标题行（缺少起息日/借方发生额/贷方发生额列）"],
            )

        # 解析元数据（标题行之前的行）
        metadata = self._parse_metadata(ws, header_row, col_map)

        # 解析数据行
        transactions, errors = self._parse_data_rows(ws, header_row + 1, col_map)

        wb.close()

        transactions.sort(key=lambda t: t.date)

        opening_balance = self._parse_amount(metadata.get("opening_balance", ""))
        closing_balance = self._parse_amount(metadata.get("closing_balance", ""))

        return ParseResult(
            transactions=transactions,
            bank=self.BANK_NAME,
            statement_date=transactions[-1].date if transactions else None,
            opening_balance=opening_balance,
            closing_balance=closing_balance,
            confidence=self.confidence,
            errors=errors,
        )

    def _find_header(self, ws) -> Tuple[Optional[int], Dict[int, str]]:
        """定位标题行，返回 (行号, {列号: 标准化字段名})。

        扫描所有行，找到同时包含"借方发生额"和"贷方发生额"的行作为标题行。
        """
        import openpyxl

        for row_idx in range(1, min(ws.max_row + 1, 50)):
            row_values = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=row_idx, column=c).value
                if v is not None:
                    row_values[c] = str(v).strip()

            row_text = "".join(row_values.values())

            # 必须同时包含这三个关键字
            if all(kw in row_text for kw in self.HEADER_KEYWORDS):
                col_map = {}
                for c, val in row_values.items():
                    mapped = self.COLUMN_MAP.get(val)
                    if mapped:
                        col_map[c] = mapped
                return row_idx, col_map

        return None, {}

    def _parse_metadata(self, ws, header_row: int, col_map: Dict[int, str]) -> Dict[str, str]:
        """解析标题行之前的元数据行，提取账号、期初/期末余额等。"""
        metadata: Dict[str, str] = {}

        for r in range(1, header_row):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                text = str(v).strip()
                if any(k in text for k in ("对账单期初余额", "期初余额", "上期余额")):
                    bal = self._find_neighbor_number(ws, r, c)
                    if bal:
                        metadata["opening_balance"] = bal
                elif any(k in text for k in ("对账单余额", "期末余额")):
                    bal = self._find_neighbor_number(ws, r, c)
                    if bal:
                        metadata["closing_balance"] = bal
                elif "账号名称" in text or "户名" in text:
                    name = self._find_neighbor_text(ws, r, c)
                    if name:
                        metadata["account_name"] = name
                elif text == "账号" or text.startswith("账号"):
                    acct = self._find_neighbor_text(ws, r, c)
                    # 排除"账号名称"
                    if acct and "名称" not in acct:
                        metadata["account_no"] = acct

        return metadata

    def _parse_data_rows(self, ws, start_row: int, col_map: Dict[int, str]) -> Tuple[List[Transaction], List[str]]:
        transactions = []
        errors = []

        for r in range(start_row, ws.max_row + 1):
            row_data = {}
            for c, field in col_map.items():
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    row_data[field] = str(v).strip()

            if not row_data:
                continue

            # 必须至少有一个金额字段
            if "debit_amount" not in row_data and "credit_amount" not in row_data:
                continue

            try:
                txn = self._row_to_transaction(row_data)
                if txn:
                    transactions.append(txn)
            except Exception as e:
                errors.append(f"Row {r}: {e}")
                self.confidence = max(0, self.confidence - 0.01)

        return transactions, errors

    def _row_to_transaction(self, row: Dict[str, str]) -> Optional[Transaction]:
        date_str = row.get("value_date") or row.get("trade_date") or ""
        tx_date = self._parse_date(date_str)
        if not tx_date:
            return None

        debit = self._parse_amount(row.get("debit_amount", ""))
        credit = self._parse_amount(row.get("credit_amount", ""))

        if debit and debit > 0:
            amount = debit
            direction = "expense"
        elif credit and credit > 0:
            amount = credit
            direction = "income"
        else:
            return None

        description = row.get("summary", "")
        if not description:
            trans_type = row.get("transaction_type", "")
            description = trans_type or ("支出交易" if direction == "expense" else "收入交易")

        reference_number = row.get("serial_no") or None
        counterparty = row.get("counterparty") or None

        notes_parts = []
        balance = self._parse_amount(row.get("balance", ""))
        if balance is not None:
            notes_parts.append(f"余额:{balance}")
        cparty_acct = row.get("counterparty_account", "")
        if cparty_acct:
            notes_parts.append(f"对方账号:{cparty_acct}")
        business_ref = row.get("business_ref", "")
        if business_ref and business_ref != reference_number:
            notes_parts.append(f"业务参考号:{business_ref}")
        notes = "; ".join(notes_parts) if notes_parts else None

        account_no = row.get("account_no") or None
        account_name = row.get("account_name") or None

        return Transaction(
            date=tx_date,
            description=description,
            amount=amount,
            currency="CNY",
            direction=direction,
            counterparty=counterparty,
            reference_number=reference_number,
            notes=notes,
            account_number=account_no,
            account_name=account_name,
        )

    @staticmethod
    def _parse_date(text: str) -> Optional[date]:
        text = text.strip()
        for fmt in ["%Y-%m-%d", "%Y%m%d"]:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return None

    @staticmethod
    def _parse_amount(text: str) -> Optional[Decimal]:
        text = text.strip().replace(",", "").replace(" ", "").replace("￥", "").replace("CNY", "").replace("cny", "")
        if not text:
            return None
        try:
            return Decimal(text)
        except (InvalidOperation, Exception):
            return None

    @staticmethod
    def _find_neighbor_number(ws, row: int, col: int) -> Optional[str]:
        """在同行的右边列或下方行查找数值。"""
        # 同行右边
        for c in range(col + 1, min(col + 5, ws.max_column + 1)):
            v = ws.cell(row=row, column=c).value
            if v is not None:
                text = str(v).strip()
                if re.match(r"^[\d,]+\.?\d*$", text):
                    return text
        # 下一行同列
        return None

    @staticmethod
    def _find_neighbor_text(ws, row: int, col: int) -> Optional[str]:
        """在同行的右边列查找非空文本。"""
        for c in range(col + 1, min(col + 5, ws.max_column + 1)):
            v = ws.cell(row=row, column=c).value
            if v is not None:
                text = str(v).strip()
                if text and not text.startswith("None"):
                    return text
        return None
