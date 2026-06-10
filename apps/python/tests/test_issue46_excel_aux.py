"""Tracer bullet: 验证 Excel 导出时部门列输出 aux_category（Issue #46 补漏）。

RED 阶段：测试应 FAIL，因为 excel_builder 尚未写入 aux_category。
"""
import sys
import os
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from finance_agent_backend.tools.excel_builder import ExcelBuilder


def _make_entries():
    return [
        {
            "date": "2024-01-15", "voucher_no": 1, "entry_seq": 1,
            "summary": "支付启胜物业1月份物业费",
            "subject_code": "5060203", "subject_name": "管理费用_物业管理费",
            "debit_amount": 500.0, "credit_amount": None,
            "direction": "expense", "counterparty": "启胜物业",
            "match_source": "rule", "rule_id": "rule_e001",
            "original_summary": "支付启胜物业1月份物业费",
            "original_amount": 500.0, "is_manual": False,
            "aux_category": "04", "aux_category_name": "公共部门",
        },
        {
            "date": "2024-01-15", "voucher_no": 1, "entry_seq": 2,
            "summary": "支付启胜物业1月份物业费",
            "subject_code": "1000201", "subject_name": "工行",
            "debit_amount": None, "credit_amount": 500.0,
            "direction": "bank", "counterparty": "",
            "match_source": "auto", "rule_id": "",
            "original_summary": "", "original_amount": 0.0, "is_manual": False,
            "aux_category": "", "aux_category_name": "",
        },
    ]


def test_excel_department_column_has_aux_category():
    """Excel 第 15 列（部门）应输出 aux_category 值。"""
    entries = _make_entries()
    builder = ExcelBuilder()
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        tmp_path = f.name
    builder.build_voucher_from_entries(entries, output_path=tmp_path, period="202401")

    import openpyxl
    wb = openpyxl.load_workbook(tmp_path)
    ws = wb.active
    os.unlink(tmp_path)

    # 第 1 行是对方科目分录（row 2），第 15 列 = O 列 = 部门
    cell = ws.cell(row=2, column=15)
    assert cell.value == "04", (
        f"期望部门列='04'，实际 '{cell.value}'"
    )

    # 第 2 行是银行分录，aux_category 为空（openpyxl 空字符串 → None）
    cell_bank = ws.cell(row=3, column=15)
    assert cell_bank.value is None, (
        f"银行分录部门列应为空，实际 '{cell_bank.value}'"
    )
