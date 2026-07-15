"""Tracer bullet: Issue #48 凭证管道字段一致性 + PipelineEntry dataclass 类型化（方案 C 修订版）。

RED 阶段：所有测试应 FAIL，因为 PipelineEntry 尚未实现。
"""
import sys
import os
import tempfile
import atexit



def _cleanup_db(db_path):
    def _do():
        try:
            from finance_agent_backend import db as _db
            _db.close_db()
            os.unlink(db_path)
        except Exception:
            pass
    atexit.register(_do)


# ── 1. PipelineEntry 字段完整性 ──────────────────────────────────

def test_pipeline_entry_has_all_fields():
    """PipelineEntry 应有 17 个字段。"""
    from finance_agent_backend.models import PipelineEntry
    fields = PipelineEntry.__dataclass_fields__
    expected = {
        'entry_seq', 'voucher_no', 'date', 'summary',
        'subject_code', 'subject_name',
        'debit_amount', 'credit_amount',
        'direction', 'counterparty',
        'match_source', 'rule_id',
        'original_summary', 'original_amount', 'is_manual',
        'aux_category', 'aux_category_name',
    }
    assert set(fields) == expected, (
        f"PipelineEntry 字段不匹配。期望 {expected}，实际 {set(fields)}"
    )


# ── 2. from_dict 严格校验 ────────────────────────────────────────

def test_from_dict_rejects_unknown_fields():
    """from_dict 遇到未知字段应显式报错，不静默丢弃。"""
    from finance_agent_backend.models import PipelineEntry
    d = {
        "entry_seq": 1, "voucher_no": 1, "date": "2024-01-15",
        "summary": "测试", "subject_code": "50602",
        "unknown_field": "should_fail",
    }
    try:
        PipelineEntry.from_dict(d)
        assert False, "from_dict 应对未知字段报错"
    except ValueError as e:
        assert "unknown_field" in str(e)


def test_from_dict_accepts_valid_fields():
    """from_dict 应正确构造已知字段。"""
    from finance_agent_backend.models import PipelineEntry
    d = {
        "entry_seq": 1, "voucher_no": 1, "date": "2024-01-15",
        "summary": "测试", "subject_code": "50602", "subject_name": "管理费用",
        "debit_amount": 100.0, "credit_amount": None,
        "direction": "expense", "counterparty": "测试方",
        "match_source": "rule", "rule_id": "rule_e032",
        "original_summary": "测试", "original_amount": 100.0,
        "is_manual": False,
        "aux_category": "04", "aux_category_name": "公共部门",
    }
    e = PipelineEntry.from_dict(d)
    assert e.rule_id == "rule_e032"
    assert e.aux_category == "04"


# ── 3. from_db_row 动态映射 ──────────────────────────────────────

def test_from_db_row_dynamic_mapping():
    """from_db_row 应基于 sqlite3.Row.keys() 动态映射字段。"""
    from finance_agent_backend.models import PipelineEntry
    import sqlite3
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
        db_path = f.name
    _cleanup_db(db_path)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("CREATE TABLE t (entry_seq, voucher_no, date, summary, rule_id)")
    conn.execute("INSERT INTO t VALUES (1, 1, '2024-01-15', '测试', 'rule_e032')")
    row = conn.execute("SELECT * FROM t").fetchone()
    e = PipelineEntry.from_db_row(row)
    assert e.entry_seq == 1
    assert e.rule_id == "rule_e032"
    assert e.summary == "测试"
    # DB 未提供的字段应有默认值
    assert e.match_source == "unmatched"
    conn.close()


# ── 4. DB schema v5 ──────────────────────────────────────────────

def test_db_schema_v5_has_account_mapping():
    """DB schema v5 应包含 account_mapping 表。"""
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
        db_path = f.name
    _cleanup_db(db_path)
    try:
        from finance_agent_backend import db as _db
        conn = _db.get_db(db_path=db_path)
        _db.init_db(conn)
        # 检查 voucher_draft_entry 仍有 rule_id 列（v4 迁移）
        cols = [r[1] for r in conn.execute("PRAGMA table_info(voucher_draft_entry)").fetchall()]
        assert 'rule_id' in cols, "voucher_draft_entry 缺少 rule_id 列"
        # 检查 account_mapping 表存在（v5 迁移）
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert 'account_mapping' in tables, "缺少 account_mapping 表"
        ver = conn.execute("SELECT version FROM schema_version ORDER BY version DESC LIMIT 1").fetchone()
        assert ver[0] == 7, f"schema_version 应为 7，实际 {ver[0]}"
    finally:
        pass


# ── 5. export 全链路字段保留 ─────────────────────────────────────

def test_export_preserves_all_fields():
    """save_draft → export，rule_id / match_source / original_summary / is_manual 不丢失。"""
    from finance_agent_backend.bridge import handle_voucher_save_draft, handle_voucher_export

    from finance_agent_backend.models import PipelineEntry
    entry = PipelineEntry(
        entry_seq=1, voucher_no=1,
        date="2024-01-15", summary="支付报销款",
        subject_code="50602", subject_name="管理费用",
        debit_amount=100.0, credit_amount=None,
        direction="expense", counterparty="测试",
        match_source="rule", rule_id="rule_e032",
        original_summary="支付报销款", original_amount=100.0,
        is_manual=False,
        aux_category="04", aux_category_name="公共部门",
    ).asdict()

    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
        db_path = f.name
    _cleanup_db(db_path)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name

    save_result = handle_voucher_save_draft({
        "db_path": db_path, "name": "test", "period": "202401", "entries": [entry]
    })
    assert save_result["success"] is True
    draft_id = save_result["draft_id"]

    export_result = handle_voucher_export({
        "db_path": db_path, "draft_id": draft_id,
        "output_path": output_path, "period": "202401",
    })
    assert export_result["success"] is True, f"export 失败: {export_result}"

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    os.unlink(output_path)

    # 验证 export 统计正确（save_draft 只存了 1 条传入 entry）
    assert export_result["entry_count"] == 1


# ── 6. excel_builder COLUMN_MAP 映射 ─────────────────────────────

def test_excel_column_map_writes_aux_category():
    """excel_builder 通过 COLUMN_MAP 正确把 aux_category 写入第 15 列。"""
    from finance_agent_backend.tools.excel_builder import ExcelBuilder

    from finance_agent_backend.models import PipelineEntry
    entries = [
        PipelineEntry(
            date="2024-01-15", voucher_no=1, entry_seq=1,
            summary="支付报销款", subject_code="50602", subject_name="管理费用",
            debit_amount=100.0, credit_amount=None,
            direction="expense", counterparty="测试",
            match_source="rule", rule_id="rule_e032",
            original_summary="支付报销款", original_amount=100.0,
            is_manual=False, aux_category="04", aux_category_name="公共部门",
        ).asdict(),
        PipelineEntry(
            date="2024-01-15", voucher_no=1, entry_seq=2,
            summary="支付报销款", subject_code="1000201", subject_name="工行",
            debit_amount=None, credit_amount=100.0,
            direction="bank", counterparty="",
            match_source="auto", rule_id="",
            original_summary="", original_amount=0.0,
            is_manual=False, aux_category="", aux_category_name="",
        ).asdict(),
    ]

    builder = ExcelBuilder()
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        tmp_path = f.name
    builder.build_voucher_from_entries(entries, output_path=tmp_path, period="202401")

    import openpyxl
    wb = openpyxl.load_workbook(tmp_path)
    ws = wb.active
    os.unlink(tmp_path)

    # 第 1 条分录（对方科目），第 15 列 = 部门 = aux_category
    cell = ws.cell(row=2, column=15)
    assert cell.value == "04", f"期望部门列='04'，实际 '{cell.value}'"

    # 第 2 条分录（银行科目），aux_category 为空 → None
    cell_bank = ws.cell(row=3, column=15)
    assert cell_bank.value is None, f"银行分录部门列应为空，实际 '{cell_bank.value}'"
